import streamlit as st
import os
from dotenv import load_dotenv
import google.generativeai as genai
import PyPDF2
import json
import pandas as pd
from pathlib import Path
import tempfile
import asyncio
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
import io

# Load environment variables
load_dotenv()

# Configure Gemini API
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# Gemini model configuration
PARSING_CONFIG = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    "response_mime_type": "application/json",
}

MERGING_CONFIG = {
    "temperature": 0.7,
    "top_p": 0.95,
    "top_k": 64,
    "max_output_tokens": 65536,
    "response_mime_type": "text/plain",
}

def initialize_gemini_model(model_name="gemini-2.0-flash", for_merging=False):
    """Initialize and return the Gemini model"""
    return genai.GenerativeModel(
        model_name=model_name,
        generation_config=MERGING_CONFIG if for_merging else PARSING_CONFIG,
    )

def split_pdf_to_pages(pdf_file):
    """Split PDF into individual pages and return their paths"""
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    total_pages = len(pdf_reader.pages)
    page_paths = []
    
    # Create temporary directory for pages
    temp_dir = tempfile.mkdtemp()
    
    for page_num in range(total_pages):
        # Create a new PDF writer for this page
        pdf_writer = PyPDF2.PdfWriter()
        pdf_writer.add_page(pdf_reader.pages[page_num])
        
        # Save the page to a temporary file
        page_path = os.path.join(temp_dir, f'page_{page_num + 1}.pdf')
        with open(page_path, 'wb') as output_file:
            pdf_writer.write(output_file)
        page_paths.append(page_path)
    
    return page_paths, total_pages

def create_progress_bar(progress, status, percentage):
    """Create a progress bar with percentage and status"""
    col1, col2, col3 = st.columns([2, 6, 2])
    with col1:
        st.markdown(f'<div class="progress-percentage">0%</div>', unsafe_allow_html=True)
    with col2:
        progress_bar = st.progress(0)
    with col3:
        st.markdown(f'<div class="progress-percentage">100%</div>', unsafe_allow_html=True)
    
    # Update progress and show current percentage
    progress_bar.progress(progress, f"{status} ({percentage:.0f}%)")
    return progress_bar

async def process_page(page_path, model, page_num, progress_bar, status_text):
    """Process a single PDF page"""
    try:
        start_time = time.time()
        
        # Upload to Gemini
        gemini_file = genai.upload_file(page_path, mime_type="application/pdf")
        
        # Create base prompt
        base_prompt = "Parse this PDF page into structured JSON format. Include all relevant information."
        
        # Add custom instructions if provided
        if st.session_state.custom_prompt:
            prompt = f"{base_prompt}\n\nAdditional Instructions:\n{st.session_state.custom_prompt}"
        else:
            prompt = base_prompt
        
        # Start chat session
        chat = model.start_chat()
        
        # Show that we're starting this page with timestamp
        status_text.text(f"🔄 Started page {page_num + 1} at {time.strftime('%H:%M:%S')}")
        
        response = chat.send_message([gemini_file, prompt])
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Update progress
        if progress_bar and status_text:
            current = getattr(progress_bar, 'current_count', 0) + 1
            setattr(progress_bar, 'current_count', current)
            progress_bar.progress(current / progress_bar.total_pages)
            status_text.text(f"✅ Page {page_num + 1} done in {processing_time:.1f}s (Total: {current}/{progress_bar.total_pages})")
        
        return json.loads(response.text)
    except Exception as e:
        st.error(f"Error processing page {page_num + 1}: {str(e)}")
        return None

async def process_pages_parallel(page_paths, model, progress_bar, status_text):
    """Process multiple pages in parallel using separate sessions"""
    print(f"\n[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Starting parallel processing for {len(page_paths)} pages")
    
    # Initialize progress tracking
    setattr(progress_bar, 'current_count', 0)
    progress_bar.total_pages = len(page_paths)
    
    # Pre-initialize all files and sessions
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Beginning session initialization")
    progress_bar.progress(0.1, f"Initializing... (10%)")
    sessions = []
    
    # Initialize all sessions first
    for page_num, page_path in enumerate(page_paths, 1):
        try:
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Initializing session for page {page_num}")
            # Upload file and create chat session
            gemini_file = genai.upload_file(page_path, mime_type="application/pdf")
            chat = model.start_chat()
            sessions.append((gemini_file, chat, page_path))
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Session initialized for page {page_num}")
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Error initializing session for page {page_num}: {str(e)}")
            st.error(f"Error initializing session for page {page_num}: {str(e)}")
    
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] All sessions initialized")
    progress_bar.progress(0.2, f"Processing... (20%)")
    
    # Create base prompt
    base_prompt = """Parse this PDF page into structured JSON format. Follow these strict guidelines:
1. Ensure all text values are properly escaped JSON strings
2. Remove any line breaks or special characters within text values
3. Use simple data types: strings, numbers, or boolean values only
4. Avoid nested objects deeper than 2 levels
5. If a field might contain quotes or special characters, properly escape them
6. Format all dates as ISO strings (YYYY-MM-DD)
7. Return only valid JSON that can be parsed by standard JSON parsers

Expected format:
{
    "field1": "value1",
    "field2": 123,
    "field3": "2024-03-21",
    "table_data": [
        {"column1": "value1", "column2": "value2"}
    ]
}"""
    if st.session_state.custom_prompt:
        prompt = f"{base_prompt}\n\nAdditional Instructions:\n{st.session_state.custom_prompt}"
    else:
        prompt = base_prompt
    
    # Process function that runs in a separate thread
    def process_single_page(session_data, prompt, page_num):
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Thread starting processing for page {page_num}")
        gemini_file, chat, _ = session_data
        response = chat.send_message([gemini_file, prompt])
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Thread completed processing for page {page_num}")
        return response.text
    
    # Async wrapper for processing
    async def process_with_session(session_data, page_num):
        try:
            # Log the exact start time
            current_time = datetime.now().strftime("%H:%M:%S.%f")[:-3]
            print(f"[{current_time}] Starting async task for page {page_num + 1}")
            
            # Run the API call in a thread pool
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Dispatching page {page_num + 1} to thread pool")
            response_text = await asyncio.to_thread(process_single_page, session_data, prompt, page_num + 1)
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Thread pool completed for page {page_num + 1}")
            
            # Update progress
            current = getattr(progress_bar, 'current_count', 0) + 1
            setattr(progress_bar, 'current_count', current)
            progress_percent = 0.2 + (0.6 * current / progress_bar.total_pages)
            percentage = progress_percent * 100
            progress_bar.progress(progress_percent, f"Processing... ({percentage:.0f}%)")
            
            return json.loads(response_text)
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Error processing page {page_num + 1}: {str(e)}")
            st.error(f"Error processing page {page_num + 1}: {str(e)}")
            return None
    
    # Create all tasks
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Creating all tasks")
    
    # Create and start all tasks at once
    tasks = [asyncio.create_task(process_with_session(session, i)) for i, session in enumerate(sessions)]
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] Created {len(tasks)} tasks, waiting for completion")
    
    # Wait for all tasks to complete
    results = await asyncio.gather(*tasks)
    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] All tasks completed")
    
    # Final progress update
    progress_bar.progress(0.8, "Filtering... (80%)")
    
    return results

async def post_merge_validation(merged_json, model):
    """Additional validation and consolidation of merged tables using Gemini thinking model"""
    try:
        # Convert merged JSON to string
        json_str = json.dumps(merged_json, indent=2)
        print("\n[DEBUG] Starting post-merge validation")
        
        # Create prompt for post-merge validation
        prompt = f"""
        Analyze and optimize this JSON structure by consolidating similar tables and removing insignificant data.
        Follow these guidelines strictly:

        1. Table Consolidation Rules:
           - Combine tables with similar names or purposes into a single table
           - Merge tables that share similar column structures
           - Remove redundant "Sr No" or "S.No" columns and add a new sequential numbering
           - Keep unique identifiers and test parameters intact

        2. Table Quality and Filtering Rules:
           - Remove tables that have only 1-2 rows unless they contain critical metadata
           - Filter out tables that lack meaningful column structure
           - Remove tables where more than 50% of cells are empty or null
           - Exclude tables that don't provide significant analytical value
           - Keep tables only if they contain actual tabular data (not just key-value pairs)
           - Remove any row where only the first column has data and all other columns are empty/null
           - Remove rows that are used only for section headers or subtotals

        3. Similarity Criteria:
           - Tables containing "Test Results" should be merged into one comprehensive table
           - Tables with 70% or more matching columns should be consolidated
           - Tables with similar prefixes or suffixes should be evaluated for merging
           - If a table's content can be merged into another more comprehensive table, do so

        4. Data Structure Requirements:
           - Maintain data integrity during consolidation
           - Ensure consistent column naming across merged tables
           - Add a new "sequence_number" column starting from 0 for each table
           - Remove any duplicate entries
           - Standardize date formats and numerical values
           - Ensure all retained tables have proper headers and consistent data types
           - For merged tables, ensure sequence numbers are continuous starting from 0

        5. Return the optimized data in this exact format:
        {{
            "data": [
                {{
                    "table": "Consolidated Table Name",
                    "rows": [
                        {{"sequence_number": 0, "parameter": "value1", "result": "value2", ...}},
                        {{"sequence_number": 1, "parameter": "value3", "result": "value4", ...}}
                    ]
                }}
            ]
        }}

        Important:
        - Only keep tables that provide meaningful analytical value
        - Ensure each table has sufficient rows to justify its existence
        - Merge similar tables to create more comprehensive datasets
        - Remove or merge tables that contain redundant or sparse information
        - All sequence numbers must start from 0 and be continuous
        - Remove any row where only the first column contains data
        - Ensure no empty or header-only rows remain in the final output

        Input JSON to optimize:
        {json_str}
        """
        
        # Get optimized result from Gemini using the thinking model
        validation_model = initialize_gemini_model("gemini-2.0-flash-thinking-exp-01-21", for_merging=True)
        chat = validation_model.start_chat()
        response = chat.send_message(prompt)
        
        print("\n[DEBUG] Post-merge validation response:")
        print(response.text)
        
        try:
            # Extract JSON from the response text
            json_text = response.text
            if "```json" in json_text and "```" in json_text:
                json_text = json_text.split("```json")[1].split("```")[0].strip()
            
            # Parse and return the optimized JSON
            optimized_json = json.loads(json_text)
            print("[DEBUG] Successfully parsed optimized JSON")
            return optimized_json
        except Exception as e:
            print(f"[DEBUG] Error parsing optimized JSON: {str(e)}")
            print("[DEBUG] Response text that caused error:")
            print(response.text)
            raise
            
    except Exception as e:
        print(f"[DEBUG] Final error in post_merge_validation: {str(e)}")
        st.error(f"Error in post-merge validation: {str(e)}")
        return None

async def merge_with_gemini(results, model):
    """Merge JSON results using Gemini API"""
    try:
        # Debug: Log the raw results
        print("\n[DEBUG] Raw results before merging:")
        for i, result in enumerate(results, 1):
            print(f"\n[DEBUG] Page {i} result type: {type(result)}")
            print(f"[DEBUG] Page {i} result content:")
            print(result)
            if result is None:
                print(f"[DEBUG] Warning: Page {i} returned None")
                continue
            try:
                # Verify each result can be serialized
                json.dumps(result)
                print(f"[DEBUG] Page {i} is valid JSON")
            except Exception as e:
                print(f"[DEBUG] Page {i} has invalid JSON: {str(e)}")

        # Filter out None values and attempt to clean results
        cleaned_results = []
        for i, result in enumerate(results, 1):
            if result is not None:
                try:
                    if isinstance(result, str):
                        # Try to parse if it's a string
                        parsed = json.loads(result)
                        cleaned_results.append(parsed)
                        print(f"[DEBUG] Page {i}: Successfully parsed string to JSON")
                    else:
                        # If it's already a dict/list, verify it can be serialized
                        json.dumps(result)
                        cleaned_results.append(result)
                        print(f"[DEBUG] Page {i}: Valid JSON object")
                except Exception as e:
                    print(f"[DEBUG] Page {i}: Skipped due to error: {str(e)}")
                    continue

        # Convert cleaned results to a formatted string
        try:
            json_str = json.dumps(cleaned_results, indent=2)
            print("\n[DEBUG] Successfully serialized cleaned results")
        except Exception as e:
            print(f"[DEBUG] Error serializing cleaned results: {str(e)}")
            raise

        # Create prompt for merging
        prompt = f"""
        Merge these JSON results from different pages into a single coherent JSON structure.
        Follow these guidelines strictly:
        1. Combine similar tables into a single table:
           - Tables with similar names should be merged (e.g., "Test Results - Part 1", "Test Results - Part 2")
           - Remove part numbers or suffixes when merging table names
           - Combine all rows from similar tables while maintaining data integrity
           - Ensure consistent column names across merged rows

        2. Each final table should have:
           - A clear, consolidated table name (e.g., "Test Results (Pesticides Residues)")
           - Consistent column names across all rows
           - Simple data types (strings, numbers) for values
           - All data from the original tables preserved

        3. Return the data in this exact format:
        {{
            "data": [
                {{
                    "table": "Consolidated Table Name",
                    "rows": [
                        {{"column1": "value1", "column2": "value2", ...}},
                        {{"column1": "value3", "column2": "value4", ...}}
                    ]
                }}
            ]
        }}

        Important:
        - Merge tables that contain "Test Results" in their names into a single table
        - Keep Quality Characteristics and other distinct tables separate
        - Ensure no data is lost during merging

        Input JSON array to merge:
        {json_str}
        """
        
        # Get merged result from Gemini using the thinking model
        merge_model = initialize_gemini_model("gemini-2.0-flash-thinking-exp-01-21", for_merging=True)
        chat = merge_model.start_chat()
        response = chat.send_message(prompt)
        
        print("\n[DEBUG] Gemini merge response:")
        print(response.text)
        
        try:
            # Extract JSON from the response text
            json_text = response.text
            if "```json" in json_text and "```" in json_text:
                json_text = json_text.split("```json")[1].split("```")[0].strip()
            
            # Parse the merged JSON
            merged_json = json.loads(json_text)
            print("[DEBUG] Successfully parsed merged JSON")
            
            # Add post-merge validation
            print("[DEBUG] Starting post-merge validation")
            optimized_json = await post_merge_validation(merged_json, model)
            
            # Return optimized JSON if successful, otherwise return merged JSON
            return optimized_json if optimized_json is not None else merged_json
            
        except Exception as e:
            print(f"[DEBUG] Error in merge processing: {str(e)}")
            print("[DEBUG] Response text that caused error:")
            print(response.text)
            raise
            
    except Exception as e:
        print(f"[DEBUG] Final error in merge_with_gemini: {str(e)}")
        st.error(f"Error merging results: {str(e)}")
        return None

def convert_to_csv(json_data):
    """Convert JSON data to CSV format"""
    try:
        # Initialize empty DataFrame
        df = None
        
        # Case 1: If we have a samples list directly
        if isinstance(json_data, dict) and 'samples' in json_data:
            samples = json_data['samples']
            if isinstance(samples, list):
                # Convert each sample to a dictionary if it's a string
                processed_samples = []
                for sample in samples:
                    if isinstance(sample, str):
                        # Try to safely evaluate the string to a dictionary
                        try:
                            # Remove any single quotes around dictionary keys
                            sample = sample.replace("'", '"')
                            sample_dict = json.loads(sample)
                            processed_samples.append(sample_dict)
                        except:
                            continue
                    elif isinstance(sample, dict):
                        processed_samples.append(sample)
                
                if processed_samples:
                    df = pd.DataFrame(processed_samples)
        
        # Case 2: If we have a data array with tables
        elif isinstance(json_data, dict) and 'data' in json_data:
            all_rows = []
            for table in json_data['data']:
                table_name = table.get('table', '')
                rows = table.get('rows', [])
                
                # Add table name to each row
                for row in rows:
                    row['Table'] = table_name
                    all_rows.append(row)
            
            if all_rows:
                df = pd.DataFrame(all_rows)
                # Reorder columns to put Table first
                if 'Table' in df.columns:
                    cols = ['Table'] + [col for col in df.columns if col != 'Table']
                    df = df[cols]
        
        # Case 3: If we have a direct list of samples
        elif isinstance(json_data, list):
            processed_items = []
            for item in json_data:
                if isinstance(item, str):
                    try:
                        # Try to safely evaluate the string to a dictionary
                        item = item.replace("'", '"')
                        item_dict = json.loads(item)
                        processed_items.append(item_dict)
                    except:
                        continue
                elif isinstance(item, dict):
                    processed_items.append(item)
            
            if processed_items:
                df = pd.DataFrame(processed_items)
        
        if df is not None:
            return df
        else:
            st.error("Could not convert the JSON data to a proper tabular format")
            return None
            
    except Exception as e:
        st.error(f"Error converting to CSV: {str(e)}")
        return None

def cleanup_temp_files(file_paths):
    """Clean up temporary files"""
    for path in file_paths:
        try:
            os.unlink(path)
        except Exception:
            pass
    try:
        os.rmdir(os.path.dirname(file_paths[0]))
    except Exception:
        pass

def convert_to_excel(json_data):
    """Convert JSON data to Excel format with multiple sheets"""
    try:
        # Create a new workbook
        workbook = Workbook()
        
        if isinstance(json_data, dict) and 'data' in json_data:
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # Process each table into its own sheet
            for table in json_data['data']:
                table_name = table.get('table', 'Sheet')
                # Create safe sheet name (Excel has 31 char limit and some invalid chars)
                safe_name = str(table_name)[:31].replace('/', '_').replace('\\', '_')
                
                # Create new sheet
                sheet = workbook.create_sheet(title=safe_name)
                
                # Get rows for this table
                rows = table.get('rows', [])
                if rows:
                    # Write headers
                    headers = list(rows[0].keys())
                    headers.remove('Table') if 'Table' in headers else None
                    for col, header in enumerate(headers, 1):
                        sheet.cell(row=1, column=col, value=header)
                    
                    # Write data
                    for row_idx, row in enumerate(rows, 2):
                        for col_idx, header in enumerate(headers, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=row.get(header))
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    except Exception as e:
        st.error(f"Error converting to Excel: {str(e)}")
        return None

def load_saved_settings():
    """Load saved settings from local file"""
    settings_file = Path("settings.json")
    if settings_file.exists():
        try:
            with open(settings_file, "r") as f:
                return json.load(f)
        except Exception:
            return {"api_key": "", "custom_prompt": ""}
    return {"api_key": "", "custom_prompt": ""}

def save_settings(api_key="", custom_prompt=""):
    """Save settings to local file"""
    settings_file = Path("settings.json")
    try:
        settings = {
            "api_key": api_key,
            "custom_prompt": custom_prompt
        }
        with open(settings_file, "w") as f:
            json.dump(settings, f)
    except Exception as e:
        st.error(f"Error saving settings: {str(e)}")

def clear_settings(clear_api=False, clear_prompt=False):
    """Clear specified settings and session state"""
    try:
        settings_file = Path("settings.json")
        current_settings = {}
        
        # Load current settings if file exists
        if settings_file.exists():
            try:
                with open(settings_file, "r") as f:
                    current_settings = json.load(f)
            except:
                pass
        
        # Update settings based on what needs to be cleared
        if clear_api:
            current_settings["api_key"] = ""
            st.session_state.api_key = ""
        if clear_prompt:
            current_settings["custom_prompt"] = ""
            st.session_state.custom_prompt = ""
        
        # Save updated settings
        with open(settings_file, "w") as f:
            json.dump(current_settings, f)
        
        return True
    except Exception as e:
        st.error(f"Error clearing settings: {str(e)}")
        return False

def main():
    st.set_page_config(page_title="AgNext X Gemini Flash 2.0", layout="wide")
    
    # Load saved settings
    saved_settings = load_saved_settings()
    
    # Initialize session state for storing results, API key, and custom prompt
    if 'merged_json' not in st.session_state:
        st.session_state.merged_json = None
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'api_key' not in st.session_state:
        st.session_state.api_key = saved_settings.get("api_key", "") or os.getenv("GEMINI_API_KEY", "")
    if 'custom_prompt' not in st.session_state:
        st.session_state.custom_prompt = saved_settings.get("custom_prompt", "")
    if 'show_tabs' not in st.session_state:
        st.session_state.show_tabs = False
    if 'current_tab' not in st.session_state:
        st.session_state.current_tab = "JSON"
    
    # Custom CSS for title styling
    st.markdown("""
        <style>
        .title {
            text-align: center;
            color: #2E4053;
            padding: 2rem;
            border-radius: 5px;
            margin-bottom: 3.5rem;
            background: linear-gradient(to right, #f8f9fa, #ffffff, #f8f9fa);
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
        }
        .stButton > button {
            margin: 0 !important;
            padding: 0 1rem !important;
            border: none !important;
            background: transparent !important;
            color: #0066cc !important;
        }
        .stButton > button:hover {
            color: #003366 !important;
            background: #f0f0f0 !important;
        }
        .input-container {
            display: flex;
            align-items: center;
            width: 100%;
            position: relative;
        }
        .input-container .stButton {
            position: absolute;
            right: 10px;
            top: 50%;
            transform: translateY(-50%);
            z-index: 1;
        }
        .input-container input, .input-container textarea {
            width: 100% !important;
            padding-right: 40px !important;
        }
        .upload-text {
            margin-bottom: 2rem;
            padding: 1rem;
            text-align: center;
            color: #5D6D7E;
        }
        div[data-testid="stToolbar"] {
            display: none;
        }
        .data-container {
            margin-top: 2rem;
            padding: 1rem;
            border-radius: 5px;
        }
        .api-key-container, .prompt-container {
            margin: 2rem 0;
            padding: 1rem;
            border-radius: 5px;
            background: #f8f9fa;
        }
        .hint-text {
            font-size: 0.9em;
            color: #666;
            font-style: italic;
        }
        /* Hide "press enter to apply" messages */
        small {
            display: none !important;
        }
        /* Custom input container styles */
        .custom-input-container {
            position: relative;
            width: 100%;
        }
        .custom-input-container input,
        .custom-input-container textarea {
            width: 100% !important;
            padding-right: 40px !important;
            box-sizing: border-box !important;
        }
        .custom-input-container .check-button {
            position: absolute;
            right: 10px;
            top: 50%;
            transform: translateY(-50%);
            background: none;
            border: none;
            color: #0066cc;
            cursor: pointer;
            padding: 5px;
            font-size: 16px;
            z-index: 100;
        }
        .custom-input-container .check-button:hover {
            color: #003366;
        }
        /* Input field styling */
        .stTextInput > div > div > input {
            padding-right: 45px !important;
        }
        .stTextArea > div > div > textarea {
            padding-right: 45px !important;
        }
        /* Button positioning */
        .stButton > button {
            position: absolute !important;
            right: 5px !important;
            top: 50% !important;
            transform: translateY(-50%) !important;
            z-index: 1 !important;
            background: transparent !important;
            border: none !important;
            color: #0066cc !important;
            min-height: 0 !important;
            padding: 0 10px !important;
            margin: 0 !important;
            line-height: normal !important;
        }
        .stButton > button:hover {
            color: #003366 !important;
            background: transparent !important;
        }
        /* Container for input + button */
        div[data-testid="column"] {
            position: relative !important;
        }
        /* Process PDF button styling */
        div[data-testid="stButton"] > button {
            background: linear-gradient(to right, #0066cc, #0099ff) !important;
            color: white !important;
            padding: 0.75rem 2rem !important;
            border-radius: 10px !important;
            border: none !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1) !important;
            transition: all 0.3s ease !important;
            font-weight: 500 !important;
            margin: 1.5rem 0 !important;
        }
        div[data-testid="stButton"] > button:hover {
            background: linear-gradient(to right, #005bb7, #0088e8) !important;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2) !important;
            color: white !important;
        }
        div[data-testid="stButton"] > button:active {
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1) !important;
        }
        .file-info {
            background: #f8f9fa !important;
            padding: 1rem !important;
            border-radius: 8px !important;
            border: 1px solid #e9ecef !important;
            margin-bottom: 1rem !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
        }
        /* Results container spacing */
        .results-spacing {
            margin-bottom: 2rem;
            padding: 1rem;
        }
        
        /* Tab container spacing */
        .stTabs {
            margin-top: 0rem;
        }
        
        /* Process button bottom margin */
        div[data-testid="stButton"] button[kind="primary"] {
            margin-bottom: 3rem !important;
        }
        
        /* Progress bar container styling */
        .stProgress {
            margin-top: 3rem !important;
            margin-bottom: 3rem !important;
        }
        .stProgress > div > div > div {
            background-color: #d7ebff !important;
        }
        /* Progress label container */
        .progress-container {
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 0.5rem;
        }
        .progress-label {
            font-size: 0.9rem;
            color: #666;
        }
        .progress-percentage {
            font-size: 0.9rem;
            color: #0066cc;
            font-weight: 500;
        }
        </style>
        <h1 class="title">AgNext PDF Parser</h1>
    """, unsafe_allow_html=True)
    
    st.markdown('<p class="upload-text">Powered by Gemini 2.0 Flash</p>', unsafe_allow_html=True)
    
    # API Key input section with expander
    with st.expander("🔑 Add Gemini API Key"):
        container = st.container()
        with container:
            col1, col2, col3 = st.columns([18, 1, 1])
            with col1:
                api_key_input = st.text_input(
                    "Enter API Key",
                    value=st.session_state.api_key if st.session_state.api_key else "",
                    type="password",
                    placeholder="Enter your Gemini API key (optional)",
                    label_visibility="collapsed"
                )
            with col2:
                if st.button("✓", key="api_key_check", help="Apply API Key"):
                    st.session_state.api_key = api_key_input
                    save_settings(
                        api_key=api_key_input,
                        custom_prompt=st.session_state.custom_prompt
                    )
            with col3:
                if st.button("🗑️", key="clear_api_key", help="Clear API Key"):
                    if clear_settings(clear_api=True):
                        st.success("API Key cleared!")
                        st.rerun()

    # Custom prompt input section without expander
    container = st.container()
    with container:
        col1, col2, col3 = st.columns([18, 1, 1])
        with col1:
            custom_prompt = st.text_area(
                "Additional Instructions",
                value=st.session_state.custom_prompt,
                placeholder="Enter any specific instructions for parsing (optional)",
                label_visibility="collapsed"
            )
        with col2:
            if st.button("✓", key="apply_instructions", help="Apply Instructions"):
                st.session_state.custom_prompt = custom_prompt
                save_settings(
                    api_key=st.session_state.api_key,
                    custom_prompt=custom_prompt
                )
        with col3:
            if st.button("🗑️", key="clear_instructions", help="Clear Instructions"):
                if clear_settings(clear_prompt=True):
                    st.success("Instructions cleared!")
                    st.rerun()

    # Add checkbox for multi-sheet option before file upload
    use_multi_sheet = st.checkbox("Split different tables into separate sheets", 
                                help="When checked, each different table type will be exported to a separate sheet in the Excel file")

    # Initialize model with appropriate API key
    if st.session_state.api_key:
        genai.configure(api_key=st.session_state.api_key)
        model = initialize_gemini_model()
    else:
        env_api_key = os.getenv("GEMINI_API_KEY")
        if not env_api_key:
            st.error("⚠️ No API key found. Please either add it in the configuration above or set it in the environment variables.")
            st.stop()
        genai.configure(api_key=env_api_key)
        model = initialize_gemini_model()
    
    # File uploader
    uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")
    
    if uploaded_file is not None:
        # Store file info in session state
        if 'uploaded_file_info' not in st.session_state or st.session_state.uploaded_file_info['name'] != uploaded_file.name:
            st.session_state.uploaded_file_info = {
                'name': uploaded_file.name,
                'size': uploaded_file.size,
                'type': uploaded_file.type
            }
            # Clear results only when new file is uploaded
            st.session_state.merged_json = None
            st.session_state.df = None
        
        # Read PDF info
        pdf_reader = PyPDF2.PdfReader(uploaded_file)
        total_pages = len(pdf_reader.pages)
        
        # Display file info in a nice format
        st.markdown("### 📁 File Details")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown('<div class="file-info">Filename: ' + uploaded_file.name + '</div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="file-info">File size: {uploaded_file.size / 1024:.2f} KB</div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="file-info">Total pages: {total_pages}</div>', unsafe_allow_html=True)
        
        # Process button
        if st.button("Process PDF", use_container_width=True, type="primary") or st.session_state.merged_json is not None:
            if st.session_state.merged_json is None:
                # Create progress bar with percentage display
                progress_bar = create_progress_bar(0, "Starting", 0)
                progress_bar.total_pages = total_pages
                status_text = st.empty()
                
                page_paths, total_pages = split_pdf_to_pages(uploaded_file)
                
                try:
                    # Process pages in parallel with separate sessions
                    results = asyncio.run(process_pages_parallel(page_paths, model, progress_bar, status_text))
                    
                    # Merge results using Gemini
                    progress_bar.progress(0.9, "Filtering... (90%)")
                    st.session_state.merged_json = asyncio.run(merge_with_gemini(results, model))
                    if st.session_state.merged_json:
                        st.session_state.df = convert_to_csv(st.session_state.merged_json)
                    progress_bar.progress(1.0, "Complete! (100%)")
                    time.sleep(1)  # Show "Complete!" for a moment
                
                finally:
                    # Clean up temporary files
                    cleanup_temp_files(page_paths)
                    progress_bar.empty()
            
            # Display results if available
            if st.session_state.merged_json:
                # Add vertical spacing
                st.markdown("<div class='results-spacing'></div>", unsafe_allow_html=True)
                
                # Create tabs for JSON and Spreadsheet views
                results_container = st.container()
                with results_container:
                    # Add section header with spacing
                    st.markdown("### 📊 Results")
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # Create tabs
                    tab1, tab2 = st.tabs(["📊 JSON View", "📈 Spreadsheet View"])
                    
                    # JSON Tab
                    with tab1:
                        st.json(st.session_state.merged_json)
                        col1, col2 = st.columns([6, 4])
                        with col1:
                            st.download_button(
                                label="⬇️ Download JSON",
                                data=json.dumps(st.session_state.merged_json, indent=2),
                                file_name="parsed_data.json",
                                mime="application/json",
                                use_container_width=True,
                                key="json_download"
                            )
                    
                    # Spreadsheet Tab
                    with tab2:
                        if use_multi_sheet:
                            # Convert to Excel with multiple sheets
                            excel_buffer = convert_to_excel(st.session_state.merged_json)
                            if excel_buffer:
                                # Create separate dataframes for each table
                                table_dfs = {}
                                if isinstance(st.session_state.merged_json, dict) and 'data' in st.session_state.merged_json:
                                    for table in st.session_state.merged_json['data']:
                                        table_name = table.get('table', 'Sheet')
                                        rows = table.get('rows', [])
                                        if rows:
                                            # Create DataFrame for this table
                                            df = pd.DataFrame(rows)
                                            # Remove 'Table' column if it exists
                                            if 'Table' in df.columns:
                                                df = df.drop('Table', axis=1)
                                            table_dfs[table_name] = df
                                
                                if table_dfs:
                                    # Create a selector for different tables
                                    st.markdown("### 📑 Select Table to View")
                                    selected_table = st.selectbox(
                                        "Choose a table to view",
                                        options=list(table_dfs.keys()),
                                        label_visibility="collapsed"
                                    )
                                    
                                    # Show the selected table
                                    st.markdown(f"**{selected_table}**")
                                    st.dataframe(
                                        table_dfs[selected_table],
                                        use_container_width=True,
                                        height=400
                                    )
                                
                                col1, col2 = st.columns([6, 4])
                                with col1:
                                    st.download_button(
                                        label="⬇️ Download Excel",
                                        data=excel_buffer,
                                        file_name="parsed_data.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        key="excel_download"
                                    )
                        else:
                            # Use existing CSV conversion
                            if st.session_state.df is not None:
                                st.dataframe(
                                    st.session_state.df,
                                    use_container_width=True,
                                    height=400
                                )
                                col1, col2 = st.columns([6, 4])
                                with col1:
                                    st.download_button(
                                        label="⬇️ Download CSV",
                                        data=st.session_state.df.to_csv(index=False),
                                        file_name="parsed_data.csv",
                                        mime="text/csv",
                                        use_container_width=True,
                                        key="csv_download"
                                    )
    else:
        # Clear session state when no file is uploaded
        if 'uploaded_file_info' in st.session_state:
            del st.session_state.uploaded_file_info
        st.session_state.merged_json = None
        st.session_state.df = None
        st.session_state.show_tabs = False

if __name__ == "__main__":
    main() 